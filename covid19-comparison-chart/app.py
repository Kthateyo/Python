import sys, getopt, time
import openpyxl as xl
from data import getJson, dateRange, Date, getPopulation, getCSV
from utils import adjustWidth, fillWithArray, getAbsolutePath

### Flags and Variables
_days_limit = 0
_starting_count = 100
_type = ''
_file = ''
_perMillion = False

### Function Definitions

# Align country data, that it starts with starting_count value
def alignFrom(country, countryName):
    start = Date(22, 1, 20)
    end = Date(29,4,20)

    # if _type == 'confirmed':
    #     start = Date(22, 1, 2020)
    #     end = Date(29,4,2020)

    countryPopulation = getPopulation(countryName)
    countryDays = []
    i = 0
    for date in dateRange(start, end):
        cases = int(country[date])
        if cases > _starting_count:
            if _perMillion == True:
                countryDays.append( cases / (countryPopulation / 1000000) )
            else:
                countryDays.append(cases)

            i += 1
            
            if _days_limit > 0:
                if i >= _days_limit:
                    break
            else:
                pass

    return countryDays


# Add to spreadsheet data of countries
def addCountries(countries):
    # get json of some countries
    data = getJson(_type)
    i = 1
    for country in countries:
        print('                                     ', end='\r')
        print(str(i)+'/'+str(len(countries)), country, end='\r')
        countryData = data[country]
        # align their data to start from 100 confirmed cases
        countryDays = alignFrom(countryData, country)
        ws.cell(1, i, country)
        fillWithArray(ws, 2, i, countryDays)
        i += 1


def drawChart(sheet):
    values = xl.chart.Reference(sheet, 1, 1, sheet.max_column, sheet.max_row)
    chart = xl.chart.LineChart()
    chart.width = 40
    chart.height = 17
    chart.add_data(values, titles_from_data=True)

    t_type = str(_type)
    t_type = t_type[0].upper() + t_type[1:]
    if t_type[-1] == 's':
        t_type = t_type[:-1]
    
    if _perMillion:
        per_million = ' per million citizens'
    else:
        per_million = ''

    chart.title = t_type + ' cases aligned from first ' + str(_starting_count)+', '+ per_million + ' for each country'
    chart.y_axis.title = 'Number of ' + str(t_type).lower() + ' cases' + per_million
    chart.x_axis.title = 'Day'
    sheet.add_chart(chart, str(sheet.cell(1, sheet.max_column + 3).column_letter) + '2')


# Prints help for usage of program
def printHelp():
    print('')
    print('That console program takes data type and list of countries and produces')
    print('a \'covid19 %Y-%m-%d %H-%M-%S.xlsx\' spreadsheet with a chart comparising those countries.')
    print('')
    print('Usage:')
    print('  python ./app.py <type> [country0, country1, ..]')
    print('')
    print('Examples:')
    print('  python ./app.py confirmed Czechia Poland')
    print('  python ./app.py deaths Germany France "Korea, South"')
    print('  python ./app.py recovered --per-million Russia China')
    print('  python ./app.py confirmed --per-million --starting-count 200 Lithuania Belarus')
    print('')
    print('type:')
    print('  confirmed      Data about confirmed cases')
    print('  deaths         Data about deaths')
    print('  recovered      Data about recovered cases')
    print('  countries      Lists available countries')
    print('')
    print('arguments:')
    print('  --days-limit <number>      Limits number of days')
    print('  --starting-count <number>  Number of cases when chart starts with for each country')
    print('  --per-million              Calculates choosen data per million citizens for each country')
    print('')


# Returns list of countries
def getCountries():
    data = getJson(_type)
    countries = []
    for country in data:
        countries.append(country)
    return countries


# Print available countries
def printCoutries():
    countries = getCountries()
    for country in countries:
        print(country)


def checkCountries(args):
    countries = getCountries()
    for arg in args:
        if not arg in countries:
            print('')
            print('\033[1m'+'Propably wrong name of country. Check if you misspelled a name with \'countries\' command'+'\033[0m')
            print('')
            exit()


### Program

## Check arguments
# Check if arguments are valid, if not print help and terminate app
args = sys.argv
if len(args) == 1 or not args[1] in ['confirmed', 'deaths', 'recovered', 'countries']:
    printHelp()
    exit()

# Set up global variable
_type = args[1]

# List available countries
if _type == 'countries':
    printCoutries()
    exit()

# Set up global variable
_file = 'data/time_series_covid19_'+ _type +'_global.csv'

# Get flag arguments
opts, countries = getopt.getopt(args[2:], 'd:s:m', ['days-limit=', 'starting-count=', 'per-million'])
for opt, arg in opts:
    if opt == '--days-limit' or opt == '-d':
        _days_limit = int(arg)
    elif opt == '--starting-count' or opt == '-s':
        _starting_count = int(arg)
    elif opt == '--per-million' or opt == '-m':
        _perMillion = True

args = countries

# Check for misspelled names
checkCountries(countries)


# create workbook
wb = xl.Workbook()
ws = wb.active

# Get countries from arguments
countries = []
for arg in args:
    countries.append(arg)

# add data to worksheet
addCountries(countries)

# create comparison chart
drawChart(ws)

# Adjust width
adjustWidth(ws)

# save workbook
name = 'covid19 ' + str((time.strftime("%Y-%m-%d %H-%M-%S"))) + '.xlsx'
path = getAbsolutePath(name)
wb.save(path)

# Print Success
print('Succesfully generated!!')
print(name)